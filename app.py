import os
import hashlib
import sqlite3
import tempfile
from datetime import datetime

import streamlit as st
import openpyxl
import pandas as pd

# ==================== CONFIGURAÇÃO ====================
st.set_page_config(
    page_title="Bolsas e Auxílios - UFPB",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

DATABASE = os.path.join(os.path.dirname(__file__), 'bolsas.db')
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

MESES = [
    (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
    (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
    (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro')
]
MESES_CURTO = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']


# ==================== CSS CUSTOMIZADO ====================
st.markdown("""
<style>
    /* Geral */
    .main .block-container { padding-top: 1rem; max-width: 100%; }
    section[data-testid="stSidebar"] > div { padding-top: 0; }

    /* Sidebar estilo Flask */
    [data-testid="stSidebar"] {
        background: #f0f4f8;
        border-right: 1px solid #dce4ec;
    }
    .sidebar-brand {
        text-align: center;
        padding: 20px 10px 15px;
        border-bottom: 1px solid #dce4ec;
        background: white;
        margin: -1rem -1rem 0 -1rem;
        padding: 20px;
    }
    .sidebar-brand h3 { color: #3a7ab5; font-weight: 700; margin: 5px 0 2px; font-size: 1.1rem; }
    .sidebar-brand small { color: #8a9bae; font-size: 0.75rem; }
    .sidebar-section {
        color: #8a9bae;
        font-size: 0.68rem;
        text-transform: uppercase;
        letter-spacing: 1px;
        font-weight: 600;
        padding: 15px 0 5px;
        margin-top: 5px;
    }
    .sidebar-link {
        display: block;
        padding: 10px 15px;
        color: #5a6d7e;
        text-decoration: none;
        font-size: 0.88rem;
        border-left: 3px solid transparent;
        border-radius: 0 6px 6px 0;
        margin: 1px 0;
        cursor: pointer;
        transition: all 0.15s;
    }
    .sidebar-link:hover { background: #e4ecf4; color: #3a7ab5; border-left-color: #6baed6; text-decoration: none; }
    .sidebar-link.active { background: #dde8f2; color: #3a7ab5; border-left-color: #4a90c4; font-weight: 600; }
    .sidebar-link i { width: 22px; display: inline-block; text-align: center; margin-right: 8px; color: #4a90c4; }
    .sidebar-user {
        background: white;
        border-top: 1px solid #dce4ec;
        padding: 12px 15px;
        margin: 0 -1rem -1rem;
        position: fixed;
        bottom: 0;
        width: inherit;
    }
    .sidebar-user .name { font-weight: 600; color: #2c3e50; font-size: 0.85rem; }
    .sidebar-user .role { color: #8a9bae; font-size: 0.72rem; }

    /* Topbar */
    .topbar {
        background: white;
        border-bottom: 1px solid #e8eef3;
        padding: 10px 0;
        margin-bottom: 20px;
        display: flex;
        justify-content: space-between;
        align-items: center;
    }
    .topbar .page-title { font-size: 1.3rem; font-weight: 700; color: #2c3e50; }
    .topbar .user-info { color: #5a6d7e; font-size: 0.85rem; }

    /* Stat cards */
    .stat-box {
        background: white;
        border-radius: 10px;
        padding: 18px;
        border: 1px solid #e8eef3;
        border-left: 4px solid #4a90c4;
        box-shadow: 0 1px 4px rgba(0,0,0,0.04);
        text-align: center;
        margin-bottom: 10px;
    }
    .stat-box.green { border-left-color: #5cb85c; }
    .stat-box.orange { border-left-color: #f5a623; }
    .stat-box.blue { border-left-color: #5bc0de; }
    .stat-box .valor { font-size: 1.4rem; font-weight: 700; color: #2c3e50; }
    .stat-box .label { font-size: 0.72rem; color: #8a9bae; text-transform: uppercase; letter-spacing: 0.3px; }

    /* Planilha lançamentos */
    .planilha-container {
        overflow-x: auto;
        background: white;
        border: 1px solid #e8eef3;
        border-radius: 10px;
        box-shadow: 0 1px 4px rgba(0,0,0,0.04);
    }
    .planilha-container table {
        border-collapse: collapse;
        font-size: 0.72rem;
        width: 100%;
    }
    .planilha-container th {
        background: #e8f0f8;
        color: #2c3e50;
        padding: 7px 5px;
        border: 1px solid #d4dfe8;
        text-align: center;
        font-size: 0.68rem;
        text-transform: uppercase;
        white-space: nowrap;
        font-weight: 600;
    }
    .planilha-container td {
        padding: 5px 4px;
        border: 1px solid #e8eef3;
        text-align: right;
        white-space: nowrap;
    }
    .planilha-container .col-pi {
        text-align: left;
        background: #fafcfd;
        font-weight: 600;
        min-width: 240px;
        max-width: 300px;
        position: sticky;
        left: 0;
        z-index: 1;
        border-right: 2px solid #d4dfe8;
    }
    .planilha-container .sub-prev { background: #fef8ec; color: #b8860b; font-size: 0.6rem; }
    .planilha-container .sub-pag { background: #ecf7ec; color: #2d7a2d; font-size: 0.6rem; }
    .planilha-container .cell-prev { background: #fffcf5; color: #6b5900; }
    .planilha-container .cell-pag { background: #f5fbf5; color: #1a5c1a; }
    .planilha-container .cell-prev.has { font-weight: 600; color: #8b6914; }
    .planilha-container .cell-pag.has { font-weight: 600; color: #1e7a1e; }
    .planilha-container .cell-total-prev { background: #fdf3d7; color: #8b6914; font-weight: 700; }
    .planilha-container .cell-total-pag { background: #d4edda; color: #155724; font-weight: 700; }
    .planilha-container .cell-detalhado { background: #f2f8ee; color: #2d5a1e; font-weight: 700; border-left: 2px solid #b8d4a8; }
    .planilha-container .cell-detalhar { background: #f8f0fc; color: #5b2d8e; font-weight: 700; }
    .planilha-container .cell-dif-pag { background: #eef4fa; color: #1a4a7a; font-weight: 700; }
    .planilha-container .th-detalhado { background: #e4f0dc; color: #2d5a1e; border-left: 2px solid #b8d4a8; }
    .planilha-container .th-detalhar { background: #f0e6f6; color: #5b2d8e; }
    .planilha-container .th-dif-pag { background: #dce8f4; color: #1a4a7a; }
    .planilha-container .th-total { background: #d4e8d4; }
    .planilha-container tr:hover td { background-color: #edf3f8 !important; }
    .positivo { color: #28a745; }
    .negativo { color: #dc3545; }

    /* Legenda */
    .legenda {
        background: white;
        border: 1px solid #e8eef3;
        border-radius: 8px;
        padding: 8px 15px;
        font-size: 0.78rem;
        display: flex;
        gap: 15px;
        flex-wrap: wrap;
        align-items: center;
        margin-bottom: 12px;
    }
    .legenda span.dot {
        display: inline-block; width: 12px; height: 12px;
        border-radius: 2px; margin-right: 4px; vertical-align: middle;
    }

    /* Card container */
    .card-container {
        background: white;
        border: 1px solid #e8eef3;
        border-radius: 10px;
        padding: 20px;
        box-shadow: 0 1px 4px rgba(0,0,0,0.04);
        margin-bottom: 15px;
    }
    .card-container h4 { color: #2c3e50; font-size: 1rem; font-weight: 600; margin-bottom: 15px; }

    /* Ocultar form border padrão */
    div[data-testid="stForm"] { border: none; padding: 0; }

    /* Upload area */
    [data-testid="stFileUploader"] {
        background: #fafcfd;
        border-radius: 12px;
        padding: 10px;
    }

    /* Botão sair */
    .btn-sair {
        display: inline-block;
        background: #f8f0f0;
        border: 1px solid #e8d4d4;
        color: #dc3545;
        padding: 6px 18px;
        border-radius: 8px;
        font-size: 0.82rem;
        font-weight: 500;
        cursor: pointer;
        text-decoration: none;
        text-align: center;
        transition: all 0.2s;
    }
    .btn-sair:hover { background: #dc3545; color: white; border-color: #dc3545; }

    /* Esconde radio padrão do Streamlit e mostra como menu */
    div[data-testid="stSidebar"] div[role="radiogroup"] { display: none; }
</style>
""", unsafe_allow_html=True)


# ==================== BANCO DE DADOS ====================
def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            login TEXT UNIQUE NOT NULL,
            senha TEXT NOT NULL,
            nome TEXT NOT NULL,
            is_admin INTEGER DEFAULT 0,
            ativo INTEGER DEFAULT 1,
            criado_em TEXT DEFAULT (datetime('now', 'localtime'))
        );
        CREATE TABLE IF NOT EXISTS usuario_pi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id INTEGER NOT NULL,
            pi_codigo TEXT NOT NULL,
            FOREIGN KEY (usuario_id) REFERENCES usuarios(id),
            UNIQUE(usuario_id, pi_codigo)
        );
        CREATE TABLE IF NOT EXISTS dados_siafi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ug_codigo TEXT, ug_nome TEXT, acao_codigo TEXT, acao_descricao TEXT,
            po_codigo TEXT, po_descricao TEXT, ptres TEXT, fonte_codigo TEXT,
            fonte_descricao TEXT, pi_codigo TEXT, pi_descricao TEXT,
            nd_codigo TEXT, nd_descricao TEXT,
            credito_disponivel REAL DEFAULT 0, despesas_empenhadas REAL DEFAULT 0,
            despesas_liquidadas REAL DEFAULT 0, despesas_liquidadas_pagar REAL DEFAULT 0,
            despesas_pagas REAL DEFAULT 0, restos_pagar REAL DEFAULT 0,
            data_upload TEXT, arquivo_origem TEXT
        );
        CREATE TABLE IF NOT EXISTS uploads_siafi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_arquivo TEXT NOT NULL,
            data_upload TEXT DEFAULT (datetime('now', 'localtime')),
            usuario_id INTEGER, registros_importados INTEGER DEFAULT 0,
            FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
        );
        CREATE TABLE IF NOT EXISTS lancamentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id INTEGER NOT NULL, pi_codigo TEXT NOT NULL,
            ano INTEGER NOT NULL, mes INTEGER NOT NULL,
            valor_previsao REAL DEFAULT 0, valor_pagamento REAL DEFAULT 0,
            observacao TEXT,
            atualizado_em TEXT DEFAULT (datetime('now', 'localtime')),
            FOREIGN KEY (usuario_id) REFERENCES usuarios(id),
            UNIQUE(usuario_id, pi_codigo, ano, mes)
        );
        CREATE INDEX IF NOT EXISTS idx_siafi_pi ON dados_siafi(pi_codigo);
        CREATE INDEX IF NOT EXISTS idx_lancamentos_pi ON lancamentos(pi_codigo, ano, mes);
        CREATE INDEX IF NOT EXISTS idx_usuario_pi ON usuario_pi(usuario_id);
    ''')
    admin = conn.execute("SELECT id FROM usuarios WHERE login = 'codeor'").fetchone()
    if not admin:
        conn.execute(
            "INSERT INTO usuarios (login, senha, nome, is_admin) VALUES (?, ?, ?, 1)",
            ('codeor', hash_password('Codeor01@'), 'Administrador CODEOR')
        )
    conn.commit()
    conn.close()


init_db()


# ==================== HELPERS ====================
def fmt_brl(valor):
    if valor is None:
        valor = 0
    try:
        valor = float(valor)
    except (ValueError, TypeError):
        return "-"
    if valor == 0:
        return "-"
    return f"{valor:,.0f}".replace(",", ".")


def processar_arquivo_siafi(file_bytes, filename):
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp.write(file_bytes)
    tmp.close()

    wb = openpyxl.load_workbook(tmp.name, data_only=True)
    ws = wb[wb.sheetnames[0]]
    conn = get_db()
    conn.execute("DELETE FROM dados_siafi")

    registros = 0
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=False):
        pi_codigo = row[9].value
        if not pi_codigo:
            continue
        conn.execute('''
            INSERT INTO dados_siafi (
                ug_codigo, ug_nome, acao_codigo, acao_descricao,
                po_codigo, po_descricao, ptres, fonte_codigo, fonte_descricao,
                pi_codigo, pi_descricao, nd_codigo, nd_descricao,
                credito_disponivel, despesas_empenhadas, despesas_liquidadas,
                despesas_liquidadas_pagar, despesas_pagas, restos_pagar,
                data_upload, arquivo_origem
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now','localtime'),?)
        ''', (
            row[0].value, row[1].value, row[2].value, row[3].value,
            row[4].value, row[5].value, row[6].value, row[7].value,
            row[8].value, row[9].value, row[10].value, row[11].value,
            row[12].value,
            row[13].value or 0, row[14].value or 0, row[15].value or 0,
            row[16].value or 0, row[17].value or 0, row[18].value or 0,
            filename
        ))
        registros += 1

    conn.execute(
        "INSERT INTO uploads_siafi (nome_arquivo, usuario_id, registros_importados) VALUES (?, ?, ?)",
        (filename, st.session_state.user_id, registros)
    )
    conn.commit()
    conn.close()
    os.unlink(tmp.name)
    return registros


# ==================== AUTENTICAÇÃO ====================
def pagina_login():
    st.markdown("<br><br>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1.2, 1, 1.2])
    with col2:
        st.markdown("""
        <div style="text-align:center; margin-bottom: 25px;">
            <div style="font-size: 3rem; color: #4a90c4;">🎓</div>
            <h2 style="color: #3a7ab5; font-weight: 700; margin: 5px 0;">Bolsas e Auxílios</h2>
            <p style="color: #8a9bae; font-size: 0.9rem;">Sistema de Monitoramento - UFPB</p>
        </div>
        """, unsafe_allow_html=True)

        with st.form("login_form"):
            login = st.text_input("👤 Login", placeholder="Digite seu login")
            senha = st.text_input("🔒 Senha", type="password", placeholder="Digite sua senha")
            submit = st.form_submit_button("Entrar", use_container_width=True, type="primary")

        if submit:
            if not login or not senha:
                st.error("Preencha login e senha.")
                return
            conn = get_db()
            user = conn.execute(
                "SELECT * FROM usuarios WHERE login = ? AND senha = ? AND ativo = 1",
                (login.strip(), hash_password(senha))
            ).fetchone()
            conn.close()
            if user:
                st.session_state.user_id = user['id']
                st.session_state.user_nome = user['nome']
                st.session_state.user_login = user['login']
                st.session_state.is_admin = bool(user['is_admin'])
                st.session_state.logado = True
                st.session_state.pagina = "Dashboard"
                st.rerun()
            else:
                st.error("Login ou senha inválidos.")

        st.caption("UFPB - Coordenação de Orçamento")


# ==================== SIDEBAR ====================
def sidebar():
    with st.sidebar:
        # Brand
        st.markdown("""
        <div class="sidebar-brand">
            <div style="font-size: 1.5rem; color: #4a90c4;">🎓</div>
            <h3>Bolsas e Auxílios</h3>
            <small>UFPB - Sistema de Monitoramento</small>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("")

        # Menu items
        menu_items = [
            ("Dashboard", "📊"),
            ("Lançamentos", "📋"),
            ("Relatórios", "📈"),
        ]

        if st.session_state.is_admin:
            admin_items = [
                ("Upload SIAFI", "📤"),
                ("Dados SIAFI", "🗄️"),
                ("Gerenciar Usuários", "👥"),
            ]
        else:
            admin_items = []

        # Página atual
        if "pagina" not in st.session_state:
            st.session_state.pagina = "Dashboard"

        # Seção principal
        st.markdown('<div class="sidebar-section">Principal</div>', unsafe_allow_html=True)
        for nome, icone in menu_items:
            active = "active" if st.session_state.pagina == nome else ""
            if st.button(f"{icone}  {nome}", key=f"menu_{nome}", use_container_width=True,
                        type="primary" if active else "secondary"):
                st.session_state.pagina = nome
                st.rerun()

        # Seção admin
        if admin_items:
            st.markdown('<div class="sidebar-section">Administração</div>', unsafe_allow_html=True)
            for nome, icone in admin_items:
                active = "active" if st.session_state.pagina == nome else ""
                if st.button(f"{icone}  {nome}", key=f"menu_{nome}", use_container_width=True,
                            type="primary" if active else "secondary"):
                    st.session_state.pagina = nome
                    st.rerun()

        # User info + sair
        st.markdown("---")
        st.markdown(f"**👤 {st.session_state.user_nome}**")
        st.caption("Administrador" if st.session_state.is_admin else "Usuário")
        if st.button("🚪 Sair do Sistema", use_container_width=True):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

    return st.session_state.pagina


# ==================== DASHBOARD ====================
def pagina_dashboard():
    st.markdown("### 📊 Dashboard")
    st.markdown("---")
    conn = get_db()

    if st.session_state.is_admin:
        total_pis = conn.execute("SELECT COUNT(DISTINCT pi_codigo) FROM dados_siafi").fetchone()[0]
        total_usuarios = conn.execute("SELECT COUNT(*) FROM usuarios WHERE is_admin = 0").fetchone()[0]
        total_uploads = conn.execute("SELECT COUNT(*) FROM uploads_siafi").fetchone()[0]
        ultimo = conn.execute("SELECT data_upload FROM uploads_siafi ORDER BY id DESC LIMIT 1").fetchone()
        total_credito = conn.execute("SELECT COALESCE(SUM(credito_disponivel), 0) FROM dados_siafi").fetchone()[0]
        total_empenhado = conn.execute("SELECT COALESCE(SUM(despesas_empenhadas), 0) FROM dados_siafi").fetchone()[0]
        total_pago = conn.execute("SELECT COALESCE(SUM(despesas_pagas), 0) FROM dados_siafi").fetchone()[0]

        c1, c2, c3, c4 = st.columns(4)
        c1.markdown(f'<div class="stat-box"><div class="label">PIs Cadastrados</div><div class="valor">{total_pis}</div></div>', unsafe_allow_html=True)
        c2.markdown(f'<div class="stat-box green"><div class="label">Usuários Ativos</div><div class="valor">{total_usuarios}</div></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="stat-box orange"><div class="label">Uploads SIAFI</div><div class="valor">{total_uploads}</div></div>', unsafe_allow_html=True)
        c4.markdown(f'<div class="stat-box blue"><div class="label">Último Upload</div><div class="valor" style="font-size:0.9rem">{ultimo["data_upload"][:16] if ultimo else "Nenhum"}</div></div>', unsafe_allow_html=True)

        st.markdown("")
        c1, c2, c3 = st.columns(3)
        c1.markdown(f'<div class="stat-box"><div class="label">Crédito Disponível Total</div><div class="valor">R$ {fmt_brl(total_credito)}</div></div>', unsafe_allow_html=True)
        c2.markdown(f'<div class="stat-box orange"><div class="label">Total Empenhado</div><div class="valor">R$ {fmt_brl(total_empenhado)}</div></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="stat-box green"><div class="label">Total Pago</div><div class="valor">R$ {fmt_brl(total_pago)}</div></div>', unsafe_allow_html=True)

        st.markdown("")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown('<div class="card-container"><h4>⚡ Ações Rápidas</h4></div>', unsafe_allow_html=True)
            c1, c2 = st.columns(2)
            with c1:
                if st.button("📤 Upload SIAFI", use_container_width=True, type="primary"):
                    st.session_state.pagina = "Upload SIAFI"
                    st.rerun()
                if st.button("📋 Lançamentos", use_container_width=True):
                    st.session_state.pagina = "Lançamentos"
                    st.rerun()
            with c2:
                if st.button("👥 Gerenciar Usuários", use_container_width=True):
                    st.session_state.pagina = "Gerenciar Usuários"
                    st.rerun()
                if st.button("📈 Relatórios", use_container_width=True):
                    st.session_state.pagina = "Relatórios"
                    st.rerun()
    else:
        user_pis = conn.execute("SELECT pi_codigo FROM usuario_pi WHERE usuario_id = ?", (st.session_state.user_id,)).fetchall()
        pi_list = [p['pi_codigo'] for p in user_pis]
        meus_pis = len(pi_list)
        total_prev = total_pgto = 0
        if pi_list:
            ph = ','.join('?' * len(pi_list))
            total_prev = conn.execute(f"SELECT COALESCE(SUM(valor_previsao),0) FROM lancamentos WHERE pi_codigo IN ({ph}) AND ano=?", pi_list + [datetime.now().year]).fetchone()[0]
            total_pgto = conn.execute(f"SELECT COALESCE(SUM(valor_pagamento),0) FROM lancamentos WHERE pi_codigo IN ({ph}) AND ano=?", pi_list + [datetime.now().year]).fetchone()[0]

        c1, c2, c3 = st.columns(3)
        c1.markdown(f'<div class="stat-box"><div class="label">Meus PIs</div><div class="valor">{meus_pis}</div></div>', unsafe_allow_html=True)
        c2.markdown(f'<div class="stat-box orange"><div class="label">Total Previsão {datetime.now().year}</div><div class="valor">R$ {fmt_brl(total_prev)}</div></div>', unsafe_allow_html=True)
        c3.markdown(f'<div class="stat-box green"><div class="label">Total Pagamento {datetime.now().year}</div><div class="valor">R$ {fmt_brl(total_pgto)}</div></div>', unsafe_allow_html=True)

        st.markdown("")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("📋 Meus Lançamentos", use_container_width=True, type="primary"):
                st.session_state.pagina = "Lançamentos"
                st.rerun()
        with c2:
            if st.button("📈 Relatórios", use_container_width=True):
                st.session_state.pagina = "Relatórios"
                st.rerun()

    conn.close()


# ==================== UPLOAD SIAFI ====================
def pagina_upload_siafi():
    if st.button("🏠 Início", key="inicio_upload"):
        st.session_state.pagina = "Dashboard"
        st.rerun()
    st.markdown("### 📤 Upload Relatório SIAFI")
    st.markdown("---")

    col1, col2 = st.columns([3, 2])
    with col1:
        st.markdown('<div class="card-container">', unsafe_allow_html=True)
        arquivo = st.file_uploader(
            "Envie o arquivo 'Relatório BOLSAS E AUXÍLIOS' (.xlsx)",
            type=['xlsx', 'xls'],
            help="Arquivo exportado do SIAFI"
        )
        if arquivo:
            st.success(f"Arquivo selecionado: **{arquivo.name}**")
            c1, c2 = st.columns([3, 1])
            with c1:
                if st.button("⬆️ Atualizar Dados SIAFI", type="primary", use_container_width=True):
                    with st.spinner("Processando arquivo..."):
                        try:
                            registros = processar_arquivo_siafi(arquivo.getvalue(), arquivo.name)
                            st.success(f"Arquivo importado com sucesso! **{registros}** registros processados.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erro ao processar: {e}")
            with c2:
                if st.button("🔄 Atualizar", use_container_width=True):
                    st.rerun()

        st.markdown('</div>', unsafe_allow_html=True)
        st.info("📋 **Instruções:** Faça o upload diário do arquivo 'Relatório BOLSAS E AUXÍLIOS' exportado do SIAFI. O sistema substituirá os dados anteriores pelos novos dados do arquivo.")

    with col2:
        st.markdown('<div class="card-container"><h4>📜 Histórico de Uploads</h4>', unsafe_allow_html=True)
        conn = get_db()
        uploads = conn.execute("SELECT * FROM uploads_siafi ORDER BY id DESC LIMIT 15").fetchall()
        conn.close()
        if uploads:
            for u in uploads:
                st.markdown(f"📄 `{u['data_upload'][:16]}` — **{u['registros_importados']}** reg.")
        else:
            st.caption("Nenhum upload realizado.")
        st.markdown('</div>', unsafe_allow_html=True)


# ==================== DADOS SIAFI ====================
def pagina_dados_siafi():
    if st.button("🏠 Início", key="inicio_dados"):
        st.session_state.pagina = "Dashboard"
        st.rerun()
    st.markdown("### 🗄️ Dados SIAFI Importados")
    st.markdown("---")
    conn = get_db()
    dados = conn.execute("SELECT * FROM dados_siafi ORDER BY pi_descricao").fetchall()
    conn.close()

    if not dados:
        st.warning("Nenhum dado importado. Faça o upload do relatório SIAFI.")
        return

    col1, col2 = st.columns([3, 1])
    with col1:
        filtro = st.text_input("🔍 Filtrar por PI ou descrição", placeholder="Digite para filtrar...")
    with col2:
        st.markdown(f"<br>**{len(dados)}** registros importados", unsafe_allow_html=True)

    rows = []
    for d in dados:
        row_text = f"{d['pi_codigo']} {d['pi_descricao']}".lower()
        if filtro and filtro.lower() not in row_text:
            continue
        rows.append({
            'PI': d['pi_codigo'], 'Descrição': d['pi_descricao'],
            'ND': d['nd_codigo'], 'PTRES': d['ptres'],
            'Créd.Disp.': d['credito_disponivel'] or 0,
            'Empenhado': d['despesas_empenhadas'] or 0,
            'Liquidado': d['despesas_liquidadas'] or 0,
            'Liq. a Pagar': d['despesas_liquidadas_pagar'] or 0,
            'Pago': d['despesas_pagas'] or 0,
            'Restos Pagar': d['restos_pagar'] or 0,
        })

    if rows:
        df = pd.DataFrame(rows)
        st.dataframe(df, use_container_width=True, hide_index=True, height=500)


# ==================== GERENCIAR USUÁRIOS ====================
def pagina_usuarios():
    if st.button("🏠 Início", key="inicio_usuarios"):
        st.session_state.pagina = "Dashboard"
        st.rerun()
    st.markdown("### 👥 Gerenciar Usuários")
    st.markdown("---")
    conn = get_db()

    tab1, tab2 = st.tabs(["📋 Lista de Usuários", "➕ Novo Usuário"])

    with tab1:
        usuarios = conn.execute('''
            SELECT u.*, GROUP_CONCAT(up.pi_codigo, ', ') as pis_vinculados
            FROM usuarios u LEFT JOIN usuario_pi up ON u.id = up.usuario_id
            WHERE u.is_admin = 0 GROUP BY u.id ORDER BY u.nome
        ''').fetchall()

        if not usuarios:
            st.info("Nenhum usuário cadastrado ainda.")
        else:
            for u in usuarios:
                status = "🟢" if u['ativo'] else "🔴"
                with st.expander(f"{status} **{u['nome']}** — `{u['login']}`"):
                    c1, c2 = st.columns(2)
                    with c1:
                        novo_nome = st.text_input("Nome", value=u['nome'], key=f"nome_{u['id']}")
                        nova_senha = st.text_input("Nova senha (vazio = manter)", type="password", key=f"senha_{u['id']}")
                        ativo = st.checkbox("Ativo", value=bool(u['ativo']), key=f"ativo_{u['id']}")
                    with c2:
                        pis_disponiveis = conn.execute("SELECT DISTINCT pi_codigo, pi_descricao FROM dados_siafi ORDER BY pi_descricao").fetchall()
                        pis_usuario = [r['pi_codigo'] for r in conn.execute("SELECT pi_codigo FROM usuario_pi WHERE usuario_id=?", (u['id'],)).fetchall()]
                        opcoes = [f"{p['pi_codigo']} - {p['pi_descricao']}" for p in pis_disponiveis]
                        default = [f"{p['pi_codigo']} - {p['pi_descricao']}" for p in pis_disponiveis if p['pi_codigo'] in pis_usuario]
                        selecionados = st.multiselect("PIs vinculados", opcoes, default=default, key=f"pis_{u['id']}")

                    bc1, bc2, bc3 = st.columns([2, 2, 6])
                    with bc1:
                        if st.button("💾 Salvar", key=f"salvar_{u['id']}", type="primary", use_container_width=True):
                            conn.execute("UPDATE usuarios SET nome=?, ativo=? WHERE id=?", (novo_nome, 1 if ativo else 0, u['id']))
                            if nova_senha:
                                conn.execute("UPDATE usuarios SET senha=? WHERE id=?", (hash_password(nova_senha), u['id']))
                            conn.execute("DELETE FROM usuario_pi WHERE usuario_id=?", (u['id'],))
                            for s in selecionados:
                                pi_cod = s.split(" - ")[0]
                                conn.execute("INSERT INTO usuario_pi (usuario_id, pi_codigo) VALUES (?,?)", (u['id'], pi_cod))
                            conn.commit()
                            st.success("Usuário atualizado!")
                            st.rerun()
                    with bc2:
                        if st.button("🗑️ Excluir", key=f"excluir_{u['id']}", use_container_width=True):
                            conn.execute("DELETE FROM usuario_pi WHERE usuario_id=?", (u['id'],))
                            conn.execute("DELETE FROM lancamentos WHERE usuario_id=?", (u['id'],))
                            conn.execute("DELETE FROM usuarios WHERE id=? AND is_admin=0", (u['id'],))
                            conn.commit()
                            st.success("Usuário excluído!")
                            st.rerun()

    with tab2:
        st.markdown('<div class="card-container">', unsafe_allow_html=True)
        with st.form("form_novo_usuario"):
            c1, c2 = st.columns(2)
            with c1:
                nome = st.text_input("Nome completo *")
                login = st.text_input("Login *")
            with c2:
                senha = st.text_input("Senha *", type="password")

            pis_disp = conn.execute("SELECT DISTINCT pi_codigo, pi_descricao FROM dados_siafi ORDER BY pi_descricao").fetchall()
            opcoes_pi = [f"{p['pi_codigo']} - {p['pi_descricao']}" for p in pis_disp]
            pis_sel = st.multiselect("Vincular PIs ao usuário", opcoes_pi)

            if st.form_submit_button("✅ Criar Usuário", type="primary", use_container_width=True):
                if not nome or not login or not senha:
                    st.error("Preencha todos os campos obrigatórios.")
                else:
                    try:
                        cursor = conn.execute("INSERT INTO usuarios (login, senha, nome) VALUES (?,?,?)",
                                            (login.strip(), hash_password(senha), nome.strip()))
                        uid = cursor.lastrowid
                        for s in pis_sel:
                            pi_cod = s.split(" - ")[0]
                            conn.execute("INSERT INTO usuario_pi (usuario_id, pi_codigo) VALUES (?,?)", (uid, pi_cod))
                        conn.commit()
                        st.success(f'Usuário "{nome}" criado com sucesso!')
                        st.rerun()
                    except sqlite3.IntegrityError:
                        st.error("Login já existe. Escolha outro.")
        st.markdown('</div>', unsafe_allow_html=True)

    conn.close()


# ==================== LANÇAMENTOS (PLANILHA) ====================
def pagina_lancamentos():
    if st.button("🏠 Início", key="inicio_lancamentos"):
        st.session_state.pagina = "Dashboard"
        st.rerun()
    st.markdown("### 📋 Lançamentos - Previsão e Pagamento")
    st.markdown("---")
    conn = get_db()

    col_ano, col_filtro = st.columns([1, 3])
    with col_ano:
        ano = st.selectbox("Ano", list(range(2024, 2031)), index=list(range(2024, 2031)).index(datetime.now().year), key="ano_lanc")
    with col_filtro:
        filtro = st.text_input("🔍 Filtrar PI", placeholder="Digite para filtrar...", key="filtro_lanc")

    if st.session_state.is_admin:
        pis = conn.execute("SELECT DISTINCT pi_codigo, pi_descricao FROM dados_siafi ORDER BY pi_descricao").fetchall()
    else:
        pis = conn.execute('''
            SELECT DISTINCT ds.pi_codigo, ds.pi_descricao FROM dados_siafi ds
            INNER JOIN usuario_pi up ON ds.pi_codigo = up.pi_codigo
            WHERE up.usuario_id = ? ORDER BY ds.pi_descricao
        ''', (st.session_state.user_id,)).fetchall()

    if not pis:
        st.warning("Nenhum PI disponível." + (" Faça upload SIAFI." if st.session_state.is_admin else " Solicite vinculação ao admin."))
        conn.close()
        return

    # Buscar lançamentos
    lanc_data = {}
    for l in conn.execute("SELECT * FROM lancamentos WHERE ano=?", (ano,)).fetchall():
        lanc_data.setdefault(l['pi_codigo'], {})[l['mes']] = dict(l)

    # Buscar SIAFI por PI
    siafi_por_pi = {}
    for s in conn.execute('''
        SELECT pi_codigo, COALESCE(SUM(credito_disponivel),0) as n,
               COALESCE(SUM(despesas_empenhadas),0) as o,
               COALESCE(SUM(despesas_liquidadas),0) as p,
               COALESCE(SUM(despesas_pagas),0) as r
        FROM dados_siafi GROUP BY pi_codigo
    ''').fetchall():
        siafi_por_pi[s['pi_codigo']] = dict(s)

    conn.close()

    # Filtrar PIs
    pis_filtrados = [pi for pi in pis if not filtro or filtro.lower() in f"{pi['pi_codigo']} {pi['pi_descricao']}".lower()]

    # Seletor de PI para lançamento
    col_sel, col_btn = st.columns([5, 1])
    with col_sel:
        opcoes_pi = {f"{pi['pi_codigo']} - {pi['pi_descricao']}": pi['pi_codigo'] for pi in pis_filtrados}
        pi_sel = st.selectbox("✏️ Selecione o PI para lançar", list(opcoes_pi.keys()), key="sel_pi_lanc", label_visibility="collapsed",
                              placeholder="✏️ Selecione um PI para lançar dados...")
    with col_btn:
        if st.button("✏️ Lançar", type="primary", use_container_width=True):
            if pi_sel:
                st.session_state.pagina = "Lançar por PI"
                st.session_state.pi_selecionado = pi_sel
                st.rerun()

    # Legenda
    st.markdown("""
    <div class="legenda">
        <b style="color:#8a9bae">Legenda:</b>
        <span><span class="dot" style="background:#f5a623"></span> Previsão</span>
        <span><span class="dot" style="background:#5cb85c"></span> Pagamento</span>
        <span><span class="dot" style="background:#dce8d4"></span> Detalhado = N+O+P</span>
        <span><span class="dot" style="background:#e8d4f0"></span> A Detalhar = SIAFI(N+O+P) - Registros</span>
        <span><span class="dot" style="background:#d4e4f0"></span> Dif.Pag = Total Pag - SIAFI(R)</span>
    </div>
    """, unsafe_allow_html=True)

    # Montar tabela HTML
    html = '<div class="planilha-container"><table>'
    html += '<tr><th rowspan="2" style="min-width:220px">PI - Descrição</th>'
    for m in MESES_CURTO:
        html += f'<th colspan="2">{m}</th>'
    html += '<th colspan="2" class="th-total">Totais</th>'
    html += '<th rowspan="2" class="th-detalhado">Detalhado</th>'
    html += '<th rowspan="2" class="th-detalhar">A Detalhar</th>'
    html += '<th rowspan="2" class="th-dif-pag">Dif. Pag</th></tr>'
    html += '<tr>'
    for _ in range(12):
        html += '<th class="sub-prev">Prev</th><th class="sub-pag">Pag</th>'
    html += '<th class="sub-prev" style="background:#d4e8d4">Prev</th><th class="sub-pag" style="background:#d4e8d4">Pag</th>'
    html += '</tr>'

    for pi in pis_filtrados:
        pi_cod = pi['pi_codigo']
        pi_desc = pi['pi_descricao']
        pi_lanc = lanc_data.get(pi_cod, {})
        siafi = siafi_por_pi.get(pi_cod, {})
        total_prev = total_pag = total_efetivo = 0

        html += f'<tr><td class="col-pi"><div style="font-size:0.78rem">{pi_desc}</div><small style="color:#8a9bae;font-size:0.63rem">{pi_cod}</small></td>'

        for mes in range(1, 13):
            ml = pi_lanc.get(mes, {})
            pv = ml.get('valor_previsao', 0) or 0
            pg = ml.get('valor_pagamento', 0) or 0
            total_prev += pv
            total_pag += pg
            total_efetivo += pg if pg > 0 else pv
            cls_p = "cell-prev has" if pv > 0 else "cell-prev"
            cls_g = "cell-pag has" if pg > 0 else "cell-pag"
            html += f'<td class="{cls_p}">{fmt_brl(pv)}</td><td class="{cls_g}">{fmt_brl(pg)}</td>'

        html += f'<td class="cell-total-prev">{fmt_brl(total_prev)}</td>'
        html += f'<td class="cell-total-pag">{fmt_brl(total_pag)}</td>'

        detalhado = (siafi.get('n', 0) or 0) + (siafi.get('o', 0) or 0) + (siafi.get('p', 0) or 0)
        html += f'<td class="cell-detalhado">{fmt_brl(detalhado)}</td>'

        a_detalhar = detalhado - total_efetivo
        cls_ad = "positivo" if a_detalhar > 0 else ("negativo" if a_detalhar < 0 else "")
        html += f'<td class="cell-detalhar"><span class="{cls_ad}">{fmt_brl(a_detalhar)}</span></td>'

        dif_pag = total_pag - (siafi.get('r', 0) or 0)
        cls_dp = "positivo" if dif_pag > 0 else ("negativo" if dif_pag < 0 else "")
        html += f'<td class="cell-dif-pag"><span class="{cls_dp}">{fmt_brl(dif_pag)}</span></td>'

        html += '</tr>'

    html += '</table></div>'
    st.markdown(html, unsafe_allow_html=True)


# ==================== LANÇAR POR PI ====================
def pagina_lancar_pi():
    c1, c2, c3 = st.columns([1, 1, 8])
    with c1:
        if st.button("⬅️ Voltar", key="voltar_lanc_pi"):
            st.session_state.pagina = "Lançamentos"
            st.rerun()
    with c2:
        if st.button("🏠 Início", key="inicio_lanc_pi"):
            st.session_state.pagina = "Dashboard"
            st.rerun()
    st.markdown("### ✏️ Lançar Previsão e Pagamento por PI")
    st.markdown("---")
    conn = get_db()

    col1, col2 = st.columns([3, 1])
    with col2:
        ano = st.selectbox("Ano", list(range(2024, 2031)), index=list(range(2024, 2031)).index(datetime.now().year), key="ano_lanc_pi")

    if st.session_state.is_admin:
        pis = conn.execute("SELECT DISTINCT pi_codigo, pi_descricao FROM dados_siafi ORDER BY pi_descricao").fetchall()
    else:
        pis = conn.execute('''
            SELECT DISTINCT ds.pi_codigo, ds.pi_descricao FROM dados_siafi ds
            INNER JOIN usuario_pi up ON ds.pi_codigo = up.pi_codigo
            WHERE up.usuario_id = ? ORDER BY ds.pi_descricao
        ''', (st.session_state.user_id,)).fetchall()

    if not pis:
        st.warning("Nenhum PI disponível.")
        conn.close()
        return

    with col1:
        opcoes = {f"{p['pi_codigo']} - {p['pi_descricao']}": p['pi_codigo'] for p in pis}
        opcoes_list = list(opcoes.keys())
        # Se veio da planilha com PI pré-selecionado
        default_idx = 0
        if "pi_selecionado" in st.session_state and st.session_state.pi_selecionado in opcoes_list:
            default_idx = opcoes_list.index(st.session_state.pi_selecionado)
        sel = st.selectbox("Selecione o PI", opcoes_list, index=default_idx)
    pi_codigo = opcoes[sel]

    # Verificar acesso
    if not st.session_state.is_admin:
        acesso = conn.execute("SELECT 1 FROM usuario_pi WHERE usuario_id=? AND pi_codigo=?",
                            (st.session_state.user_id, pi_codigo)).fetchone()
        if not acesso:
            st.error("Sem acesso a este PI.")
            conn.close()
            return

    # Dados SIAFI do PI
    siafi_data = conn.execute("SELECT * FROM dados_siafi WHERE pi_codigo=?", (pi_codigo,)).fetchall()
    if siafi_data:
        st.markdown('<div class="card-container"><h4>🗄️ Dados SIAFI Atuais</h4>', unsafe_allow_html=True)
        for s in siafi_data:
            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("Créd. Disp.", fmt_brl(s['credito_disponivel']))
            c2.metric("Empenhado", fmt_brl(s['despesas_empenhadas']))
            c3.metric("Liquidado", fmt_brl(s['despesas_liquidadas']))
            c4.metric("Pago", fmt_brl(s['despesas_pagas']))
            c5.metric("Restos Pagar", fmt_brl(s['restos_pagar']))
        st.markdown('</div>', unsafe_allow_html=True)

    # Lançamentos existentes
    lanc_existentes = {}
    for r in conn.execute("SELECT * FROM lancamentos WHERE pi_codigo=? AND ano=?", (pi_codigo, ano)).fetchall():
        lanc_existentes[r['mes']] = dict(r)

    st.markdown(f"#### 📅 Lançamentos Mensais — {ano}")
    st.markdown("🟡 **Previsão** | 🟢 **Pagamento**")

    with st.form(f"form_lancamento_{pi_codigo}_{ano}"):
        valores = {}
        cols_por_linha = 4
        for i in range(0, 12, cols_por_linha):
            cols = st.columns(cols_por_linha)
            for j, col in enumerate(cols):
                mes_idx = i + j
                if mes_idx >= 12:
                    break
                mes_num = mes_idx + 1
                mes_nome = MESES[mes_idx][1]
                lanc = lanc_existentes.get(mes_num, {})
                with col:
                    st.markdown(f"**{mes_nome}**")
                    prev = st.number_input(
                        "Previsão (R$)", value=float(lanc.get('valor_previsao', 0) or 0),
                        min_value=0.0, step=100.0, format="%.2f", key=f"prev_{mes_num}"
                    )
                    pag = st.number_input(
                        "Pagamento (R$)", value=float(lanc.get('valor_pagamento', 0) or 0),
                        min_value=0.0, step=100.0, format="%.2f", key=f"pag_{mes_num}"
                    )
                    obs = st.text_input("Obs", value=lanc.get('observacao', '') or '', key=f"obs_{mes_num}")
                    valores[mes_num] = (prev, pag, obs)
            if i + cols_por_linha < 12:
                st.markdown("---")

        if st.form_submit_button("💾 Salvar Lançamentos", type="primary", use_container_width=True):
            for mes_num, (prev, pag, obs) in valores.items():
                conn.execute('''
                    INSERT INTO lancamentos (usuario_id, pi_codigo, ano, mes, valor_previsao, valor_pagamento, observacao)
                    VALUES (?,?,?,?,?,?,?)
                    ON CONFLICT(usuario_id, pi_codigo, ano, mes)
                    DO UPDATE SET valor_previsao=excluded.valor_previsao,
                        valor_pagamento=excluded.valor_pagamento,
                        observacao=excluded.observacao,
                        atualizado_em=datetime('now','localtime')
                ''', (st.session_state.user_id, pi_codigo, ano, mes_num, prev, pag, obs))
            conn.commit()
            st.success("Lançamentos salvos com sucesso!")
            st.rerun()

    conn.close()


# ==================== RELATÓRIOS ====================
def pagina_relatorios():
    if st.button("🏠 Início", key="inicio_relatorios"):
        st.session_state.pagina = "Dashboard"
        st.rerun()
    st.markdown("### 📈 Relatório Consolidado")
    st.markdown("---")
    conn = get_db()
    ano = st.selectbox("Ano", list(range(2024, 2031)), index=list(range(2024, 2031)).index(datetime.now().year), key="ano_rel")

    if st.session_state.is_admin:
        dados = conn.execute('''
            SELECT pi_codigo, pi_descricao,
                   COALESCE(SUM(credito_disponivel),0) as credito,
                   COALESCE(SUM(despesas_empenhadas),0) as empenhado,
                   COALESCE(SUM(despesas_liquidadas),0) as liquidado,
                   COALESCE(SUM(despesas_pagas),0) as pago
            FROM dados_siafi GROUP BY pi_codigo ORDER BY pi_descricao
        ''').fetchall()
    else:
        dados = conn.execute('''
            SELECT ds.pi_codigo, ds.pi_descricao,
                   COALESCE(SUM(ds.credito_disponivel),0) as credito,
                   COALESCE(SUM(ds.despesas_empenhadas),0) as empenhado,
                   COALESCE(SUM(ds.despesas_liquidadas),0) as liquidado,
                   COALESCE(SUM(ds.despesas_pagas),0) as pago
            FROM dados_siafi ds
            INNER JOIN usuario_pi up ON ds.pi_codigo = up.pi_codigo AND up.usuario_id = ?
            GROUP BY ds.pi_codigo ORDER BY ds.pi_descricao
        ''', (st.session_state.user_id,)).fetchall()

    lanc_totais = {}
    for l in conn.execute("SELECT pi_codigo, SUM(valor_previsao) as tp, SUM(valor_pagamento) as tpg FROM lancamentos WHERE ano=? GROUP BY pi_codigo", (ano,)).fetchall():
        lanc_totais[l['pi_codigo']] = (l['tp'] or 0, l['tpg'] or 0)

    conn.close()

    if not dados:
        st.warning("Nenhum dado disponível.")
        return

    rows = []
    for d in dados:
        tp, tpg = lanc_totais.get(d['pi_codigo'], (0, 0))
        rows.append({
            'PI': d['pi_codigo'], 'Descrição': d['pi_descricao'],
            'Créd.Disp.': d['credito'], 'Empenhado': d['empenhado'],
            'Liquidado': d['liquidado'], 'Pago SIAFI': d['pago'],
            'Previsão': tp, 'Pagamento': tpg, 'Diferença': tp - tpg
        })

    df = pd.DataFrame(rows)

    filtro = st.text_input("🔍 Filtrar", placeholder="Filtrar por PI ou descrição...")
    if filtro:
        mask = df.apply(lambda r: filtro.lower() in f"{r['PI']} {r['Descrição']}".lower(), axis=1)
        df = df[mask]

    st.dataframe(df, use_container_width=True, hide_index=True, height=500)

    st.markdown("---")
    c1, c2, c3 = st.columns(3)
    c1.metric("Total Previsão", f"R$ {fmt_brl(df['Previsão'].sum())}")
    c2.metric("Total Pagamento", f"R$ {fmt_brl(df['Pagamento'].sum())}")
    c3.metric("Diferença", f"R$ {fmt_brl(df['Diferença'].sum())}")


# ==================== ROTEAMENTO PRINCIPAL ====================
def main():
    if not st.session_state.get('logado'):
        pagina_login()
        return

    pagina = sidebar()

    if pagina == "Dashboard":
        pagina_dashboard()
    elif pagina == "Lançamentos":
        pagina_lancamentos()
    elif pagina == "Lançar por PI":
        pagina_lancar_pi()
    elif pagina == "Relatórios":
        pagina_relatorios()
    elif pagina == "Upload SIAFI":
        pagina_upload_siafi()
    elif pagina == "Dados SIAFI":
        pagina_dados_siafi()
    elif pagina == "Gerenciar Usuários":
        pagina_usuarios()


if __name__ == "__main__":
    main()
